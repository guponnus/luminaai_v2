from __future__ import annotations

import random
import re
from pathlib import Path
from typing import Any

import openpyxl

from .extensions import db
from .models import Dataset, PanelRecord


PANEL_TITLES = [
    ("A", "Asset Health Overview"),
    ("B", "Network Coverage Map"),
    ("C", "Risk Matrix"),
    ("D", "Maintenance Queue"),
    ("E", "RUL Analysis"),
    ("F", "Coverage Gap"),
    ("G", "Work Order Hub"),
    ("H", "CAPEX Planner"),
    ("I", "Compliance & Inspection"),
]


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _panel_code(sheet_name: str) -> str:
    normalized = sheet_name.upper()
    if "KPI" in normalized:
        return "KPI"
    pillar_match = re.match(r"F(\d{2})[_\s-]", normalized)
    if pillar_match:
        return f"F{pillar_match.group(1)}"
    parts = sheet_name.replace("—", "-").split()
    if "Panel" in parts and len(parts) > parts.index("Panel") + 1:
        return parts[parts.index("Panel") + 1].strip("-").upper()
    return sheet_name[:3].upper()


def import_workbook(path: str | Path) -> dict[str, int]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return generate_sample_data()

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    stats: dict[str, int] = {}
    workbook_sheets = set(wb.sheetnames)

    for dataset in Dataset.query.all():
        if dataset.sheet_name not in workbook_sheets:
            db.session.delete(dataset)
    db.session.flush()

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        title = str(rows[0][0] or ws.title)
        headers = [str(h).strip() for h in rows[1] if h is not None and str(h).strip()]
        if not headers:
            continue

        dataset = Dataset.query.filter_by(sheet_name=ws.title).one_or_none()
        if dataset is None:
            dataset = Dataset(sheet_name=ws.title, title=title, panel_code=_panel_code(ws.title))
            db.session.add(dataset)
            db.session.flush()
        else:
            dataset.title = title
            dataset.panel_code = _panel_code(ws.title)
            PanelRecord.query.filter_by(dataset_id=dataset.id).delete()

        count = 0
        for row_number, row in enumerate(rows[2:], start=3):
            values = list(row[: len(headers)])
            if not any(value is not None for value in values):
                continue
            data = {headers[i]: _clean_cell(values[i]) for i in range(len(headers))}
            db.session.add(PanelRecord(dataset_id=dataset.id, row_number=row_number, data=data))
            count += 1

        dataset.record_count = count
        stats[ws.title] = count

    db.session.commit()
    return stats


def generate_sample_data(records_per_panel: int = 250) -> dict[str, int]:
    random.seed(42)
    stats: dict[str, int] = {}
    zones = ["Bur Dubai", "Mirdif", "Jumeirah", "Mushrif", "Al Mamzar", "Al Warqa"]
    asset_types = ["Power Transformer", "MV Cable", "RMU", "Feeder Pillar", "Overhead Line"]

    for code, title in PANEL_TITLES:
        sheet_name = f"Panel {code} - {title}"
        dataset = Dataset.query.filter_by(sheet_name=sheet_name).one_or_none()
        if dataset is None:
            dataset = Dataset(sheet_name=sheet_name, title=title, panel_code=code)
            db.session.add(dataset)
            db.session.flush()
        else:
            PanelRecord.query.filter_by(dataset_id=dataset.id).delete()

        for idx in range(1, records_per_panel + 1):
            health = round(random.uniform(35, 98), 1)
            risk = random.randint(1, 25)
            data = {
                "Record_ID": f"{code}-{idx:04d}",
                "Asset_ID": f"AST-{random.randint(1, 9999):04d}",
                "Asset_Type": random.choice(asset_types),
                "Zone": random.choice(zones),
                "Health_Score": health,
                "Risk_Score": risk,
                "Status": "Critical" if health < 50 or risk >= 20 else "Monitor" if health < 75 else "Healthy",
                "Priority_Rank": idx,
                "Estimated_Cost_AED": random.randint(25_000, 2_500_000),
            }
            db.session.add(PanelRecord(dataset_id=dataset.id, row_number=idx + 2, data=data))

        dataset.record_count = records_per_panel
        stats[sheet_name] = records_per_panel

    db.session.commit()
    return stats
